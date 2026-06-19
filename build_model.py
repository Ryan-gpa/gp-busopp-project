"""Build NVU_Financial_Model.xlsx — Nanoveu Limited (ASX: NVU), FY ended 31 Dec 2025.
All figures in full Australian dollars. Blue = hardcoded input, black = formula.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(color="0000CC")              # hardcoded inputs
BLACK = Font(color="000000")             # formulas
BOLD = Font(bold=True, color="000000")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
TITLE = Font(bold=True, size=14, color="1F3864")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TOT_FILL = PatternFill("solid", fgColor="FCE4D6")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
top_border = Border(top=Side(style="thin", color="404040"))

CUR = '#,##0;(#,##0);"-"'               # currency, negatives in parens, zero as dash
PCT = '0.0%;(0.0%);"-"'
NUM2 = '#,##0.00;(#,##0.00);"-"'
RATIO = '0.00"x"'

wb = openpyxl.Workbook()

def style_header_row(ws, row, ncols, text_list):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
    for i, t in enumerate(text_list):
        ws.cell(row=row, column=i + 1, value=t)

def title_block(ws, title, sub):
    ws["A1"] = title; ws["A1"].font = TITLE
    ws["A2"] = sub; ws["A2"].font = Font(italic=True, size=9, color="808080")

# ---------------------------------------------------------------- P&L
ws = wb.active
ws.title = "P&L"
title_block(ws, "Nanoveu Limited — Consolidated Statement of Profit or Loss & OCI",
            "All figures in AUD. Blue = input (from filings); black = formula. H2 2025 = FY 2025 - H1 2025.")
style_header_row(ws, 4, 4, ["$", "H1 2025 (6 mths)", "H2 2025 (derived)", "FY 2025 (12 mths)"])

# rows: (label, h1, fy, kind) kind: 'in'=input line, 'sub'=subtotal formula handled below
pl = [
    ("Revenue from contracts with customers", 281115, 297582, "in"),
    ("Total revenue", None, None, "rev"),
    ("Cost of sale of goods", -104080, -929715, "in"),
    ("Gross (loss) / profit", None, None, "gp"),
    ("Other operating income", 5645, 6436, "in"),
    ("Selling and distribution expenses", -112797, -386930, "in"),
    ("Administration expenses", -1635837, -4424940, "in"),
    ("Research costs", -138137, -760818, "in"),
    ("Share based payment expense", -1332119, -1421890, "in"),
    ("Operating (loss)", None, None, "op"),
    ("Finance income", 5168, 32756, "in"),
    ("Finance costs", -6554, -13706, "in"),
    ("(Loss) before income tax", None, None, "lbt"),
    ("Income tax expense", 0, 0, "in"),
    ("(Loss) for the period / year", None, None, "net"),
    ("Other comprehensive income / (loss) - FX translation", 56359, 9462, "in"),
    ("Total comprehensive (loss)", None, None, "tci"),
    ("__sep__", None, None, "sep"),
    ("Loss attributable to non-controlling interest", -68823, -177906, "in"),
    ("Loss attributable to owners of Nanoveu Limited", -2968773, -7423319, "in"),
]
r = 5
rowmap = {}
# first pass: place input lines and remember rows
for label, h1, fy, kind in pl:
    if label == "__sep__":
        r += 1; continue
    ws.cell(row=r, column=1, value=label)
    rowmap[kind if kind in ("rev","gp","op","lbt","net","tci") else label] = r
    r += 1
# reset and write values/formulas with explicit row tracking
r = 5
detail_rows = []  # (kind, row) for line items contributing to subtotals
rows = {}
for label, h1, fy, kind in pl:
    if label == "__sep__":
        r += 1; continue
    rows[label] = r
    cell_label = ws.cell(row=r, column=1, value=label)
    if kind == "in":
        ws.cell(row=r, column=2, value=h1).font = BLUE
        ws.cell(row=r, column=4, value=fy).font = BLUE
        ws.cell(row=r, column=3, value=f"=D{r}-B{r}").font = BLACK  # H2 = FY - H1
        for col in (2,3,4):
            ws.cell(row=r, column=col).number_format = CUR
    r += 1

# Now subtotal formulas. Identify rows by label.
def rr(lbl): return rows[lbl]
for col, L in ((2,"B"),(3,"C"),(4,"D")):
    # Total revenue = revenue from contracts (single line)
    ws.cell(row=rr("Total revenue"), column=col, value=f"={L}{rr('Revenue from contracts with customers')}")
    # Gross = revenue + COGS
    ws.cell(row=rr("Gross (loss) / profit"), column=col,
            value=f"={L}{rr('Total revenue')}+{L}{rr('Cost of sale of goods')}")
    # Operating loss = GP + other income + S&D + admin + research + SBP
    ws.cell(row=rr("Operating (loss)"), column=col,
            value=(f"={L}{rr('Gross (loss) / profit')}+{L}{rr('Other operating income')}"
                   f"+{L}{rr('Selling and distribution expenses')}+{L}{rr('Administration expenses')}"
                   f"+{L}{rr('Research costs')}+{L}{rr('Share based payment expense')}"))
    # LBT = operating + finance income + finance costs
    ws.cell(row=rr("(Loss) before income tax"), column=col,
            value=f"={L}{rr('Operating (loss)')}+{L}{rr('Finance income')}+{L}{rr('Finance costs')}")
    # Net = LBT - tax
    ws.cell(row=rr("(Loss) for the period / year"), column=col,
            value=f"={L}{rr('(Loss) before income tax')}+{L}{rr('Income tax expense')}")
    # TCI = net + OCI
    ws.cell(row=rr("Total comprehensive (loss)"), column=col,
            value=f"={L}{rr('(Loss) for the period / year')}+{L}{rr('Other comprehensive income / (loss) - FX translation')}")
    for lbl in ("Total revenue","Gross (loss) / profit","Operating (loss)","(Loss) before income tax",
                "(Loss) for the period / year","Total comprehensive (loss)"):
        cc = ws.cell(row=rr(lbl), column=col); cc.number_format = CUR; cc.font = BLACK

# styling: bold subtotal labels, fills
for lbl in ("Total revenue","Gross (loss) / profit","Operating (loss)","(Loss) before income tax",
            "(Loss) for the period / year","Total comprehensive (loss)"):
    ws.cell(row=rr(lbl), column=1).font = BOLD
    for col in range(1,5):
        ws.cell(row=rr(lbl), column=col).fill = SUB_FILL
ws.cell(row=rr("(Loss) for the period / year"), column=1).font = BOLD
for col in range(1,5):
    ws.cell(row=rr("(Loss) for the period / year"), column=col).fill = TOT_FILL

PL_NET_ROW = rr("(Loss) for the period / year")
PL_REV_ROW = rr("Total revenue")
PL_GP_ROW = rr("Gross (loss) / profit")
PL_OP_ROW = rr("Operating (loss)")

ws.column_dimensions["A"].width = 52
for c in "BCD": ws.column_dimensions[c].width = 18

# ---------------------------------------------------------------- Balance Sheet
ws2 = wb.create_sheet("Balance Sheet")
title_block(ws2, "Nanoveu Limited — Consolidated Statement of Financial Position",
            "All figures in AUD. Blue = input; black = formula. Point-in-time balances (no derivation).")
style_header_row(ws2, 4, 3, ["$", "30 Jun 2025", "31 Dec 2025"])
bs = [
    ("CURRENT ASSETS", None, None, "h"),
    ("Cash and cash equivalents", 2075432, 1807712, "in"),
    ("Trade and other receivables", 34328, 11234, "in"),
    ("Prepayments", 415543, 454731, "in"),
    ("Loan receivable (current)", 0, 472374, "in"),
    ("Total Current Assets", None, None, "tca"),
    ("NON-CURRENT ASSETS", None, None, "h"),
    ("Plant and equipment", 141048, 119190, "in"),
    ("Intangible assets", 10494942, 10500307, "in"),
    ("Loan receivable (non-current)", 441902, 0, "in"),
    ("Right of use asset", 207900, 131399, "in"),
    ("Non-current prepayments", 0, 57968, "in"),
    ("Total Non-current Assets", None, None, "tnca"),
    ("TOTAL ASSETS", None, None, "ta"),
    ("CURRENT LIABILITIES", None, None, "h"),
    ("Trade and other payables", 492133, 1004988, "in"),
    ("Contract liability", 140620, 169196, "in"),
    ("Lease liability", 109937, 64168, "in"),
    ("Provisions", 8189, 17231, "in"),
    ("Loan", 0, 134011, "in"),
    ("Total Current Liabilities", None, None, "tcl"),
    ("NON-CURRENT LIABILITIES", None, None, "h"),
    ("Lease liability (long term)", 100578, 71891, "in"),
    ("Total Non-current Liabilities", None, None, "tncl"),
    ("TOTAL LIABILITIES", None, None, "tl"),
    ("NET ASSETS", None, None, "na"),
    ("EQUITY", None, None, "h"),
    ("Issued capital", 35816939, 39377691, "in"),
    ("Reserves", 1586008, 1722677, "in"),
    ("Accumulated losses", -24337829, -28792375, "in"),
    ("Total equity attributable to NVU shareholders", None, None, "teq"),
    ("Non-controlling interest", -105480, -214563, "in"),
    ("TOTAL EQUITY", None, None, "te"),
]
r = 5; brow = {}
for label, jun, dec, kind in bs:
    brow[label] = r
    cell = ws2.cell(row=r, column=1, value=label)
    if kind == "h":
        cell.font = WHITE_BOLD; cell.fill = PatternFill("solid", fgColor="8EA9DB")
        ws2.cell(row=r, column=2).fill = PatternFill("solid", fgColor="8EA9DB")
        ws2.cell(row=r, column=3).fill = PatternFill("solid", fgColor="8EA9DB")
    elif kind == "in":
        ws2.cell(row=r, column=2, value=jun).font = BLUE
        ws2.cell(row=r, column=3, value=dec).font = BLUE
        for c in (2,3): ws2.cell(row=r, column=c).number_format = CUR
    r += 1

def b(lbl): return brow[lbl]
for col, L in ((2,"B"),(3,"C")):
    ws2.cell(row=b("Total Current Assets"), column=col,
             value=f"=SUM({L}{b('Cash and cash equivalents')}:{L}{b('Loan receivable (current)')})")
    ws2.cell(row=b("Total Non-current Assets"), column=col,
             value=f"=SUM({L}{b('Plant and equipment')}:{L}{b('Non-current prepayments')})")
    ws2.cell(row=b("TOTAL ASSETS"), column=col,
             value=f"={L}{b('Total Current Assets')}+{L}{b('Total Non-current Assets')}")
    ws2.cell(row=b("Total Current Liabilities"), column=col,
             value=f"=SUM({L}{b('Trade and other payables')}:{L}{b('Loan')})")
    ws2.cell(row=b("Total Non-current Liabilities"), column=col,
             value=f"={L}{b('Lease liability (long term)')}")
    ws2.cell(row=b("TOTAL LIABILITIES"), column=col,
             value=f"={L}{b('Total Current Liabilities')}+{L}{b('Total Non-current Liabilities')}")
    ws2.cell(row=b("NET ASSETS"), column=col,
             value=f"={L}{b('TOTAL ASSETS')}-{L}{b('TOTAL LIABILITIES')}")
    ws2.cell(row=b("Total equity attributable to NVU shareholders"), column=col,
             value=f"={L}{b('Issued capital')}+{L}{b('Reserves')}+{L}{b('Accumulated losses')}")
    ws2.cell(row=b("TOTAL EQUITY"), column=col,
             value=f"={L}{b('Total equity attributable to NVU shareholders')}+{L}{b('Non-controlling interest')}")
    for lbl in ("Total Current Assets","Total Non-current Assets","TOTAL ASSETS","Total Current Liabilities",
                "Total Non-current Liabilities","TOTAL LIABILITIES","NET ASSETS",
                "Total equity attributable to NVU shareholders","TOTAL EQUITY"):
        cc = ws2.cell(row=b(lbl), column=col); cc.number_format = CUR; cc.font = BLACK
for lbl in ("Total Current Assets","Total Non-current Assets","Total Current Liabilities",
            "Total Non-current Liabilities","Total equity attributable to NVU shareholders"):
    ws2.cell(row=b(lbl), column=1).font = BOLD
    for c in range(1,4): ws2.cell(row=b(lbl), column=c).fill = SUB_FILL
for lbl in ("TOTAL ASSETS","TOTAL LIABILITIES","NET ASSETS","TOTAL EQUITY"):
    ws2.cell(row=b(lbl), column=1).font = BOLD
    for c in range(1,4): ws2.cell(row=b(lbl), column=c).fill = TOT_FILL
ws2.column_dimensions["A"].width = 46
for c in "BC": ws2.column_dimensions[c].width = 16
BS_NA_ROW = b("NET ASSETS"); BS_CA=b("Total Current Assets"); BS_CL=b("Total Current Liabilities")
BS_CASH=b("Cash and cash equivalents"); BS_TA=b("TOTAL ASSETS"); BS_INTAN=b("Intangible assets")

# ---------------------------------------------------------------- Cash Flow
ws3 = wb.create_sheet("Cash Flow")
title_block(ws3, "Nanoveu Limited — Consolidated Statement of Cash Flows",
            "All figures in AUD. Blue = input; black = formula. H2 2025 = FY 2025 - H1 2025.")
style_header_row(ws3, 4, 4, ["$", "H1 2025", "H2 2025 (derived)", "FY 2025"])
cf = [
    ("OPERATING ACTIVITIES", None, None, "h"),
    ("Receipts from customers", 131765, 161727, "in"),
    ("Payments to suppliers and employees", -2082572, -6033547, "in"),
    ("Interest received", 5168, 32756, "in"),
    ("Interest paid", -48888, -6774, "in"),
    ("Net cash (used in) operating activities", None, None, "op"),
    ("INVESTING ACTIVITIES", None, None, "h"),
    ("Payments for property, plant and equipment", 0, -76510, "in"),
    ("Cash acquired through EMASS acquisition", 2909, 2910, "in"),
    ("Loan receivable / funds loaned to third parties", -183239, -213711, "in"),
    ("Net cash (used in) investing activities", None, None, "inv"),
    ("FINANCING ACTIVITIES", None, None, "h"),
    ("Proceeds from issue of shares (net of costs)", 3916520, 7471773, "in"),
    ("Interest free loan proceeds", 0, 134011, "in"),
    ("Interest free loan repayment", -120000, -120000, "in"),
    ("Repayment of lease liabilities", -44534, -43226, "in"),
    ("Net cash provided by financing activities", None, None, "fin"),
    ("Net increase / (decrease) in cash held", None, None, "net"),
    ("Cash at the beginning of the period", 498303, 498303, "begin"),
    ("Cash at the end of the period", None, None, "end"),
]
r=5; crow={}
for label, h1, fy, kind in cf:
    crow[label]=r
    cell=ws3.cell(row=r, column=1, value=label)
    if kind=="h":
        for c in range(1,5): ws3.cell(row=r,column=c).fill=PatternFill("solid", fgColor="A9D08E")
        cell.font=WHITE_BOLD
    elif kind=="in":
        ws3.cell(row=r,column=2,value=h1).font=BLUE
        ws3.cell(row=r,column=4,value=fy).font=BLUE
        ws3.cell(row=r,column=3,value=f"=D{r}-B{r}").font=BLACK
        for c in (2,3,4): ws3.cell(row=r,column=c).number_format=CUR
    elif kind=="begin":
        # H1 begin = FY begin = 498303 (input). H2 begin = H1 end (handled after).
        ws3.cell(row=r,column=2,value=h1).font=BLUE
        ws3.cell(row=r,column=4,value=fy).font=BLUE
        for c in (2,4): ws3.cell(row=r,column=c).number_format=CUR
    r+=1
def cf_r(lbl): return crow[lbl]
for col,L in ((2,"B"),(3,"C"),(4,"D")):
    ws3.cell(row=cf_r("Net cash (used in) operating activities"),column=col,
             value=f"=SUM({L}{cf_r('Receipts from customers')}:{L}{cf_r('Interest paid')})")
    ws3.cell(row=cf_r("Net cash (used in) investing activities"),column=col,
             value=f"=SUM({L}{cf_r('Payments for property, plant and equipment')}:{L}{cf_r('Loan receivable / funds loaned to third parties')})")
    ws3.cell(row=cf_r("Net cash provided by financing activities"),column=col,
             value=f"=SUM({L}{cf_r('Proceeds from issue of shares (net of costs)')}:{L}{cf_r('Repayment of lease liabilities')})")
    ws3.cell(row=cf_r("Net increase / (decrease) in cash held"),column=col,
             value=(f"={L}{cf_r('Net cash (used in) operating activities')}+{L}{cf_r('Net cash (used in) investing activities')}"
                    f"+{L}{cf_r('Net cash provided by financing activities')}"))
    ws3.cell(row=cf_r("Cash at the end of the period"),column=col,
             value=f"={L}{cf_r('Cash at the beginning of the period')}+{L}{cf_r('Net increase / (decrease) in cash held')}")
    for lbl in ("Net cash (used in) operating activities","Net cash (used in) investing activities",
                "Net cash provided by financing activities","Net increase / (decrease) in cash held",
                "Cash at the end of the period"):
        cc=ws3.cell(row=cf_r(lbl),column=col); cc.number_format=CUR; cc.font=BLACK
# H2 beginning cash = H1 ending cash (chain)
ws3.cell(row=cf_r("Cash at the beginning of the period"),column=3,
         value=f"=B{cf_r('Cash at the end of the period')}").font=BLACK
ws3.cell(row=cf_r("Cash at the beginning of the period"),column=3).number_format=CUR
for lbl in ("Net cash (used in) operating activities","Net cash (used in) investing activities","Net cash provided by financing activities"):
    ws3.cell(row=cf_r(lbl),column=1).font=BOLD
    for c in range(1,5): ws3.cell(row=cf_r(lbl),column=c).fill=SUB_FILL
for lbl in ("Net increase / (decrease) in cash held","Cash at the end of the period"):
    ws3.cell(row=cf_r(lbl),column=1).font=BOLD
    for c in range(1,5): ws3.cell(row=cf_r(lbl),column=c).fill=TOT_FILL
# note re anomalies
note_r = cf_r("Cash at the end of the period")+2
ws3.cell(row=note_r,column=1,value="Note: H2 'Interest paid' (+42,114) and 'Repayment of lease liabilities' (+1,308) are positive because the audited FY figure is smaller in magnitude than the reviewed H1 figure — a reclassification between the interim review and the annual audit. Derived per H2 = FY - H1 as instructed.").font=Font(italic=True,size=8,color="C00000")
ws3.merge_cells(start_row=note_r,start_column=1,end_row=note_r,end_column=4)
ws3.row_dimensions[note_r].height=42
ws3.cell(row=note_r,column=1).alignment=Alignment(wrap_text=True, vertical="top")
ws3.column_dimensions["A"].width=50
for c in "BCD": ws3.column_dimensions[c].width=16
CF_OP_H1=cf_r("Net cash (used in) operating activities")

# ---------------------------------------------------------------- Segments
ws4 = wb.create_sheet("Segments")
title_block(ws4, "Nanoveu Limited — Operating Segments",
            "All figures in AUD. Blue = input; black = formula. Note: HY 2025 reported 3 product segments + Corporate; FY 2025 added an Adjustments/Eliminations column.")
# HY block
ws4["A4"]="HALF-YEAR ENDED 30 JUNE 2025"; ws4["A4"].font=Font(bold=True,size=11,color="1F3864")
style_header_row(ws4,5,6,["$","Semiconductor","EyeFly3D","Nanoshield","Corporate","Total"])
hy_seg=[
    ("External revenue",[0,268704,12410,0],"sum"),
    ("Depreciation & amortisation",[-4292,-51620,0,-19020],"sum"),
    ("Segment profit / (loss)",[-347316,-265649,-117466,-2307165],"sum"),
    ("Total assets",[10522954,1366879,0,1921263],"sum"),
    ("Total liabilities",[-16751,-474585,0,-360121],"sum"),
]
r=6
for label,vals,_ in hy_seg:
    ws4.cell(row=r,column=1,value=label).font=BOLD
    for i,v in enumerate(vals):
        cc=ws4.cell(row=r,column=2+i,value=v); cc.font=BLUE; cc.number_format=CUR
    ws4.cell(row=r,column=6,value=f"=SUM(B{r}:E{r})").number_format=CUR
    r+=1
hy_end=r-1
# FY block
ws4.cell(row=r+1,column=1,value="YEAR ENDED 31 DECEMBER 2025").font=Font(bold=True,size=11,color="1F3864")
hdr=r+2
style_header_row(ws4,hdr,8,["$","Semiconductor","EyeFly3D","Nanoshield","Corporate","Total segments","Adj/Elim","Consolidated"])
fy_seg=[
    ("External revenue",[0,285164,12418,0],0),
    ("Depreciation & amortisation",[-9442,-116120,0,-47550],0),
    ("Segment profit / (loss)",[-2500960,-1149328,-104648,-3847290],0),
    ("Total assets",[11296205,1130366,0,1560890],-432545),
    ("Total liabilities",[-595215,-274238,0,-340839],-251194),
]
r=hdr+1
for label,vals,adj in fy_seg:
    ws4.cell(row=r,column=1,value=label).font=BOLD
    for i,v in enumerate(vals):
        cc=ws4.cell(row=r,column=2+i,value=v); cc.font=BLUE; cc.number_format=CUR
    ws4.cell(row=r,column=6,value=f"=SUM(B{r}:E{r})").number_format=CUR  # total segments
    ws4.cell(row=r,column=7,value=adj).font=BLUE; ws4.cell(row=r,column=7).number_format=CUR
    ws4.cell(row=r,column=8,value=f"=F{r}+G{r}").number_format=CUR  # consolidated
    r+=1
ws4.column_dimensions["A"].width=30
for c in "BCDEFGH": ws4.column_dimensions[c].width=15

# ---------------------------------------------------------------- KPIs
ws5=wb.create_sheet("KPIs")
title_block(ws5,"Nanoveu Limited — Key Performance Indicators",
            "Black = formula linking to P&L / Balance Sheet / Cash Flow sheets. Blue = input (share counts, months).")
# inputs block
ws5["A4"]="Inputs"; ws5["A4"].font=BOLD
inp=[("Shares on issue 30 Jun 2025",746916586),
     ("Shares on issue 31 Dec 2025",987087553),
     ("Weighted avg shares H1 2025",693931856),
     ("Weighted avg shares FY 2025",825913230),
     ("Months in H1",6),("Months in FY",12)]
r=5
for lbl,v in inp:
    ws5.cell(row=r,column=1,value=lbl)
    ws5.cell(row=r,column=2,value=v).font=BLUE
    ws5.cell(row=r,column=2).number_format='#,##0'
    r+=1
ir={lbl:5+i for i,(lbl,_) in enumerate(inp)}
def I(lbl): return f"$B${ir[lbl]}"
r+=1
style_header_row(ws5,r,4,["Metric","H1 2025","H2 2025","FY 2025"])
hdrK=r; r+=1
# helpers referencing other sheets
PL=f"'P&L'"; BS=f"'Balance Sheet'"; CFs=f"'Cash Flow'"
kpis=[
    ("Gross margin %",
     f"=IF('P&L'!B{PL_REV_ROW}=0,0,'P&L'!B{PL_GP_ROW}/'P&L'!B{PL_REV_ROW})",
     f"=IF('P&L'!C{PL_REV_ROW}=0,0,'P&L'!C{PL_GP_ROW}/'P&L'!C{PL_REV_ROW})",
     f"=IF('P&L'!D{PL_REV_ROW}=0,0,'P&L'!D{PL_GP_ROW}/'P&L'!D{PL_REV_ROW})", PCT),
    ("Operating cash burn per month (AUD)",
     f"=-'Cash Flow'!B{CF_OP_H1}/{I('Months in H1')}",
     f"=-'Cash Flow'!C{CF_OP_H1}/{I('Months in H1')}",
     f"=-'Cash Flow'!D{CF_OP_H1}/{I('Months in FY')}", CUR),
    ("Cash runway (months, at period-end cash)",
     f"=IF('Cash Flow'!B{CF_OP_H1}>=0,\"n/a\",'Balance Sheet'!B{BS_CASH}/(-'Cash Flow'!B{CF_OP_H1}/{I('Months in H1')}))",
     "=\"\"",
     f"=IF('Cash Flow'!D{CF_OP_H1}>=0,\"n/a\",'Balance Sheet'!C{BS_CASH}/(-'Cash Flow'!D{CF_OP_H1}/{I('Months in FY')}))", NUM2),
    ("Net assets per share (AUD)",
     f"='Balance Sheet'!B{BS_NA_ROW}/{I('Shares on issue 30 Jun 2025')}",
     "=\"\"",
     f"='Balance Sheet'!C{BS_NA_ROW}/{I('Shares on issue 31 Dec 2025')}", '0.0000'),
    ("Net loss for the period (AUD)",
     f"='P&L'!B{PL_NET_ROW}", f"='P&L'!C{PL_NET_ROW}", f"='P&L'!D{PL_NET_ROW}", CUR),
]
for label,b1,b2,b3,fmt in kpis:
    ws5.cell(row=r,column=1,value=label)
    ws5.cell(row=r,column=2,value=b1).number_format=fmt
    ws5.cell(row=r,column=3,value=b2).number_format=fmt
    ws5.cell(row=r,column=4,value=b3).number_format=fmt
    for c in range(2,5): ws5.cell(row=r,column=c).font=BLACK
    r+=1
# revenue growth (period over prior-year period) - uses prior-year inputs
ws5.cell(row=r,column=1,value="Revenue (AUD)")
ws5.cell(row=r,column=2,value=f"='P&L'!B{PL_REV_ROW}").number_format=CUR
ws5.cell(row=r,column=4,value=f"='P&L'!D{PL_REV_ROW}").number_format=CUR
rev_now=r; r+=1
ws5.cell(row=r,column=1,value="Revenue prior corresponding period (AUD)")
ws5.cell(row=r,column=2,value=11026).font=BLUE  # H1 2024
ws5.cell(row=r,column=4,value=6873).font=BLUE   # FY 2024
for c in (2,4): ws5.cell(row=r,column=c).number_format=CUR
rev_prior=r; r+=1
ws5.cell(row=r,column=1,value="Revenue growth %")
ws5.cell(row=r,column=2,value=f"=B{rev_now}/B{rev_prior}-1").number_format=PCT
ws5.cell(row=r,column=4,value=f"=D{rev_now}/D{rev_prior}-1").number_format=PCT
for c in (2,4): ws5.cell(row=r,column=c).font=BLACK
ws5.column_dimensions["A"].width=44
for c in "BCD": ws5.column_dimensions[c].width=16
for c in range(1,5): ws5.cell(row=hdrK,column=c)

# ---------------------------------------------------------------- Ratios
ws6=wb.create_sheet("Ratios")
title_block(ws6,"Nanoveu Limited — Financial Ratios",
            "Black = formula. Liquidity at balance dates; loss per share from filings.")
style_header_row(ws6,4,3,["Ratio","30 Jun 2025 / H1","31 Dec 2025 / FY"])
ratios=[
    ("Current ratio (CA / CL)",
     f"='Balance Sheet'!B{BS_CA}/'Balance Sheet'!B{BS_CL}",
     f"='Balance Sheet'!C{BS_CA}/'Balance Sheet'!C{BS_CL}", RATIO),
    ("Cash ratio (Cash / CL)",
     f"='Balance Sheet'!B{BS_CASH}/'Balance Sheet'!B{BS_CL}",
     f"='Balance Sheet'!C{BS_CASH}/'Balance Sheet'!C{BS_CL}", RATIO),
    ("Asset turnover (Revenue / Total assets)",
     f"='P&L'!B{PL_REV_ROW}/'Balance Sheet'!B{BS_TA}",
     f"='P&L'!D{PL_REV_ROW}/'Balance Sheet'!C{BS_TA}", '0.000"x"'),
    ("Loss per share (cents) — basic & diluted", -0.44, -0.90, '0.00;(0.00);"-"'),
    ("Operating cash burn rate per month (AUD)",
     f"=-'Cash Flow'!B{CF_OP_H1}/6", f"=-'Cash Flow'!D{CF_OP_H1}/12", CUR),
]
r=5
for label,v1,v2,fmt in ratios:
    ws6.cell(row=r,column=1,value=label)
    c1=ws6.cell(row=r,column=2,value=v1); c1.number_format=fmt
    c2=ws6.cell(row=r,column=3,value=v2); c2.number_format=fmt
    if label.startswith("Loss per share"):
        c1.font=BLUE; c2.font=BLUE
    else:
        c1.font=BLACK; c2.font=BLACK
    r+=1
ws6.column_dimensions["A"].width=44
for c in "BC": ws6.column_dimensions[c].width=18

# legend on P&L
lr = PL_NET_ROW+4
ws.cell(row=lr,column=1,value="Legend:").font=BOLD
ws.cell(row=lr+1,column=1,value="Blue text = hardcoded input from NVU filings").font=BLUE
ws.cell(row=lr+2,column=1,value="Black text = formula").font=BLACK

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"

wb.save("NVU_Financial_Model.xlsx")
print("Saved NVU_Financial_Model.xlsx")
print("PL net row", PL_NET_ROW, "rev row", PL_REV_ROW, "gp row", PL_GP_ROW)
print("BS NA row", BS_NA_ROW, "CA", BS_CA, "CL", BS_CL, "cash", BS_CASH, "TA", BS_TA)
print("CF op row", CF_OP_H1)
